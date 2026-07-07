"""
excel_export.py
----------------
Generates a .xlsx class routine that visually matches the official RUET IPE
department routine template (colors, fonts, merges, layout) exactly, but
populated with a solver result (schedule) instead of the manually-typed
official routine.

Usage:
    from excel_export import build_routine_workbook
    wb = build_routine_workbook(schedule, meta)
    wb.save("routine.xlsx")            # or stream it via BytesIO
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Period / timing layout (must mirror routine.html's DAY_COLS) ───────────
DAY_COLS = [
    {"type": "period", "period": 1, "time": "8:00-8:50",   "start_min": 0},
    {"type": "period", "period": 2, "time": "8:50-9:40",   "start_min": 50},
    {"type": "period", "period": 3, "time": "9:40-10:30",  "start_min": 100},
    {"type": "break",  "period": None, "time": "10:30\n10:50", "start_min": 150},
    {"type": "period", "period": 4, "time": "10:50-11:40", "start_min": 170},
    {"type": "period", "period": 5, "time": "11:40-12:30", "start_min": 220},
    {"type": "period", "period": 6, "time": "12:30-1:20",  "start_min": 270},
    {"type": "recess", "period": None, "time": "1:20\n2:30",   "start_min": 320},
    {"type": "period", "period": 7, "time": "2:30-3:20",   "start_min": 390},
    {"type": "period", "period": 8, "time": "3:20-4:10",   "start_min": 440},
    {"type": "period", "period": 9, "time": "4:10-5:00",   "start_min": 490},
]
NCOLS = len(DAY_COLS)
MIN_TO_COL = {c["start_min"]: i for i, c in enumerate(DAY_COLS) if c["type"] == "period"}

DAY_NAMES = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday"]
DAYS_PER_BLOCK = 3   # matches the official template: 3 days on top, remaining below

# ── Colors / fonts lifted directly from the official template ──────────────
NAVY        = "FF1F3864"
MED_BLUE    = "FF2E5090"
HDR_GRAY    = "FFD6DCE4"
BREAK_BLUE  = "FFBDD7EE"
RECESS_BLUE = "FF9DC3E6"
THEORY_FILL = "FFE8F4FD"
LAB_FILL    = "FFD4EDDA"
GAP_FILL    = "FFD0D0D0"
BORDER_CLR  = "FF2C3E50"
TAGLINE_CLR = "FFBCD6EC"
DEPT_GOLD   = "FFFFDA6A"

# Year-group color cycle (label column), matches 4th/3rd/2nd/1st in template.
# Applied top-to-bottom in whatever order `groups` is given.
GROUP_COLORS = [
    ("FFBF8F00", "FFFFEECC"),  # gold      (label fill, semester-sub fill)
    ("FF375623", "FFFFEECC"),  # dark green
    ("FF1F3864", "FFFFEECC"),  # navy
    ("FFC55A11", "FFFFEECC"),  # burnt orange
]

FONT_ARIAL = "Arial"
FONT_TNR   = "Times New Roman"

thin_grey  = Side(style="thin", color="FFAAAAAA")
thin_auto  = Side(style="thin", color=BORDER_CLR)
med_navy   = Side(style="medium", color=BORDER_CLR)

BORDER_OUTER = Border(left=med_navy, right=med_navy, top=med_navy, bottom=med_navy)
BORDER_CELL  = Border(left=thin_auto, right=thin_auto, top=thin_auto, bottom=thin_auto)
BORDER_TIME  = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)


def _fill(color):
    return PatternFill(fill_type="solid", fgColor=color)


def _set(ws, coord, value, font=None, fill=None, align=None, border=None):
    c = ws[coord]
    c.value = value
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c


def build_routine_workbook(schedule, meta):
    """
    schedule: list of dicts (as returned by /routine/{job}/{sol}), each with
              day, start_min, duration, course_type, course_code, classroom,
              room_label, faculty, student_groups.
    meta:     dict with groups (list[str]), n_student_groups, DAYS, etc.
    """
    groups = meta.get("groups") or [f"Group {i+1}" for i in range(meta.get("n_student_groups", 1))]
    n_groups = len(groups)
    n_days = meta.get("DAYS", 5)
    day_names = DAY_NAMES[:n_days]

    # Build grid[group][day][col] = schedule entry (or 'SKIP' for cells
    # covered by a preceding lab's colspan), same logic as routine.html.
    grid = [[[None] * NCOLS for _ in range(n_days)] for _ in range(n_groups)]
    for e in schedule:
        if e.get("classroom", -1) == -1:
            continue
        sgs = e.get("student_groups") or []
        if not sgs:
            continue
        g, d = sgs[0], e["day"]
        if g >= n_groups or d >= n_days:
            continue
        col = MIN_TO_COL.get(e["start_min"])
        if col is None:
            continue
        grid[g][d][col] = e
        if e.get("course_type") == "lab":
            periods_left, ci = 2, col + 1
            while periods_left > 0 and ci < NCOLS:
                grid[g][d][ci] = "SKIP"
                if DAY_COLS[ci]["type"] == "period":
                    periods_left -= 1
                ci += 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Routine"
    ws.sheet_view.showGridLines = False

    n_blocks = -(-n_days // DAYS_PER_BLOCK)  # ceil
    last_col_idx = 2 + DAYS_PER_BLOCK * NCOLS   # A/B label cols + 3 days * 11 cols
    last_col_letter = get_column_letter(last_col_idx)

    # ── Title block (rows 1-4, once at the top) ─────────────────────────────
    _set(ws, "A1", "Heaven's Light is Our Guide",
         font=Font(name=FONT_ARIAL, size=9, color=TAGLINE_CLR),
         fill=_fill(NAVY), align=Alignment(horizontal="center", vertical="center"),
         border=BORDER_OUTER)
    _set(ws, "A2", "Rajshahi University of Engineering & Technology",
         font=Font(name=FONT_ARIAL, size=14, bold=True, color="FFFFFFFF"),
         fill=_fill(NAVY), align=Alignment(horizontal="center", vertical="center"),
         border=BORDER_OUTER)
    _set(ws, "A3", "Department of Industrial & Production Engineering",
         font=Font(name=FONT_ARIAL, size=11, bold=True, color=DEPT_GOLD),
         fill=_fill(NAVY), align=Alignment(horizontal="center", vertical="center"),
         border=BORDER_OUTER)
    _set(ws, "A4", "", fill=_fill(NAVY), border=BORDER_OUTER)
    for row, height in [(1, 15), (2, 20.1), (3, 15.95), (4, 14.1)]:
        ws.row_dimensions[row].height = height
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")

    row = 5
    for block in range(n_blocks):
        block_days = list(range(block * DAYS_PER_BLOCK,
                                 min((block + 1) * DAYS_PER_BLOCK, n_days)))

        day_row, period_row, time_row = row, row + 1, row + 2
        ws.row_dimensions[day_row].height = 26.25
        ws.row_dimensions[period_row].height = 12.95
        ws.row_dimensions[time_row].height = 14.1

        # "Day" / "Period" / "Class Time" label cells (col A:B merged)
        for r, label in [(day_row, "Day"), (period_row, "Period"), (time_row, "Class Time")]:
            ws.merge_cells(f"A{r}:B{r}")
            fill_c = NAVY if r == day_row else HDR_GRAY
            font_c = Font(name=FONT_ARIAL, size=9, bold=True, color="FFFFFFFF") if r == day_row \
                     else (Font(name=FONT_ARIAL, size=8, bold=True, color="FF2C3E50") if r == period_row
                           else Font(name=FONT_ARIAL, size=7, italic=True, color="FF444444"))
            _set(ws, f"A{r}", label, font=font_c, fill=_fill(fill_c),
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=BORDER_OUTER)
            ws[f"B{r}"].fill = _fill(fill_c)
            ws[f"B{r}"].border = BORDER_OUTER

        col_cursor = 3  # column C = index 3
        for d in block_days:
            start_col = col_cursor
            end_col = col_cursor + NCOLS - 1
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)

            ws.merge_cells(f"{start_letter}{day_row}:{end_letter}{day_row}")
            _set(ws, f"{start_letter}{day_row}", day_names[d],
                 font=Font(name=FONT_ARIAL, size=12, bold=True, color="FFFFFFFF"),
                 fill=_fill(MED_BLUE),
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=BORDER_OUTER)
            for ci in range(start_col, end_col + 1):
                ws[f"{get_column_letter(ci)}{day_row}"].border = BORDER_OUTER
                ws[f"{get_column_letter(ci)}{day_row}"].fill = _fill(MED_BLUE)

            for ci, colinfo in enumerate(DAY_COLS):
                letter = get_column_letter(col_cursor + ci)
                is_break, is_recess = colinfo["type"] == "break", colinfo["type"] == "recess"
                fill_c = BREAK_BLUE if is_break else RECESS_BLUE if is_recess else HDR_GRAY
                label = "Break" if is_break else "Recess" if is_recess else str(colinfo["period"])
                font_c = Font(name=FONT_ARIAL, size=7, bold=True, color="FF1F3864") if (is_break or is_recess) \
                         else Font(name=FONT_ARIAL, size=9, bold=True, color="FF000000")
                _set(ws, f"{letter}{period_row}", label, font=font_c, fill=_fill(fill_c),
                     align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                     border=BORDER_OUTER)
                _set(ws, f"{letter}{time_row}", colinfo["time"],
                     font=Font(name=FONT_ARIAL, size=6.5, italic=True, color="FF555555"),
                     fill=_fill(fill_c),
                     align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                     border=BORDER_TIME if not (is_break or is_recess) else BORDER_OUTER)

            col_cursor += NCOLS

        # ── Group rows ───────────────────────────────────────────────────
        for gi in range(n_groups):
            r = day_row + 3 + gi
            ws.row_dimensions[r].height = 42
            lbl_color, sem_color = GROUP_COLORS[gi % len(GROUP_COLORS)]

            ws.merge_cells(f"A{r}:A{r}")
            _set(ws, f"A{r}", groups[gi],
                 font=Font(name=FONT_ARIAL, size=8, bold=True, color="FFFFFFFF"),
                 fill=_fill(lbl_color),
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=BORDER_OUTER)
            _set(ws, f"B{r}", "Odd Semester",
                 font=Font(name=FONT_ARIAL, size=7, color=sem_color),
                 fill=_fill(lbl_color),
                 align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                 border=BORDER_OUTER)

            col_cursor = 3
            for d in block_days:
                skip_until = None
                for ci, colinfo in enumerate(DAY_COLS):
                    letter = get_column_letter(col_cursor + ci)
                    cell_coord = f"{letter}{r}"
                    is_break, is_recess = colinfo["type"] == "break", colinfo["type"] == "recess"

                    if is_break or is_recess:
                        fill_c = BREAK_BLUE if is_break else RECESS_BLUE
                        ws[cell_coord].fill = _fill(fill_c)
                        ws[cell_coord].border = BORDER_OUTER
                        continue

                    entry = grid[gi][d][ci]
                    if entry == "SKIP":
                        continue
                    if entry is None:
                        ws[cell_coord].border = BORDER_CELL
                        continue

                    is_lab = entry.get("course_type") == "lab"
                    code = entry.get("course_code") or f"C{entry.get('course_id')}"
                    room = entry.get("room_label") or (
                        f"Lab-{entry['classroom']+1}" if is_lab else f"R-{entry['classroom']+1:03d}"
                    )
                    fac = f"F{entry['faculty']}" if entry.get("faculty", -1) >= 0 else "TBA"
                    text = f"{code}\n{room}\n{fac}"

                    span = 1
                    if is_lab:
                        periods_found, look = 1, ci + 1
                        while periods_found < 3 and look < NCOLS:
                            span += 1
                            if DAY_COLS[look]["type"] == "period":
                                periods_found += 1
                            look += 1

                    if span > 1:
                        end_letter = get_column_letter(col_cursor + ci + span - 1)
                        ws.merge_cells(f"{letter}{r}:{end_letter}{r}")

                    _set(ws, cell_coord, text,
                         font=Font(name=FONT_TNR, size=8),
                         fill=_fill(LAB_FILL if is_lab else THEORY_FILL),
                         align=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BORDER_CELL)
                col_cursor += NCOLS

        row = day_row + 3 + n_groups
        ws.row_dimensions[row].height = 6  # spacer row like the template's row 12
        for ci in range(1, last_col_idx + 1):
            ws.cell(row=row, column=ci).fill = _fill(GAP_FILL)
        row += 1

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 9.0
    ws.column_dimensions["B"].width = 7.4
    col_cursor = 3
    for _ in range(DAYS_PER_BLOCK):
        for ci, colinfo in enumerate(DAY_COLS):
            letter = get_column_letter(col_cursor + ci)
            width = 6.5 if colinfo["type"] in ("break", "recess") else 9.4
            ws.column_dimensions[letter].width = width
        col_cursor += NCOLS

    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return wb


def workbook_to_bytes(wb) -> BytesIO:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf